import openpyxl
from django.db import transaction
from django.db.models import Count, Q
from rest_framework import mixins, status
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.viewsets import GenericViewSet
from rest_framework_simplejwt.exceptions import TokenError
from rest_framework_simplejwt.tokens import RefreshToken

from .models import User, Vehicle, Role
from .permissions import IsJefeOperacionesOrAbove, IsOperativoOrAbove
from .serializers import (
    LoginSerializer,
    UserSerializer,
    UserListSerializer,
    UserCreateSerializer,
    UserUpdateSerializer,
    VehicleSerializer,
)


class LoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data['user']
        refresh = RefreshToken.for_user(user)
        return Response({
            'access': str(refresh.access_token),
            'refresh': str(refresh),
            'user': UserSerializer(user).data,
        })


class LogoutView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            token = RefreshToken(request.data.get('refresh'))
            token.blacklist()
        except TokenError:
            pass
        return Response(status=status.HTTP_204_NO_CONTENT)


class RefreshView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        try:
            token = RefreshToken(request.data.get('refresh'))
            return Response({
                'access': str(token.access_token),
                'refresh': str(token),
            })
        except TokenError:
            return Response(
                {'detail': 'Token inválido o expirado.'},
                status=status.HTTP_401_UNAUTHORIZED,
            )


class MeView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response(UserSerializer(request.user).data)


class UserViewSet(
    mixins.ListModelMixin,
    mixins.CreateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin,
    GenericViewSet,
):
    permission_classes = [IsJefeOperacionesOrAbove]

    def get_queryset(self):
        user = self.request.user
        qs = User.objects.select_related('campus_asignado').annotate(
            vehiculos=Count('vehicles', filter=Q(vehicles__activo=True), distinct=True),
            sanciones_activas=Count('sanctions', filter=Q(sanctions__estado='activa'), distinct=True),
        )
        if user.rol != Role.RECTOR:
            qs = qs.filter(campus_asignado=user.campus_asignado)

        rol = self.request.query_params.get('rol')
        estado = self.request.query_params.get('estado')
        search = self.request.query_params.get('search')
        if rol:
            qs = qs.filter(rol=rol)
        if estado:
            qs = qs.filter(estado=estado)
        if search:
            qs = qs.filter(
                Q(nombre__icontains=search) |
                Q(apellido__icontains=search) |
                Q(codigo_institucional__icontains=search)
            )
        return qs

    def get_serializer_class(self):
        if self.action == 'create':
            return UserCreateSerializer
        if self.action in ('update', 'partial_update'):
            return UserUpdateSerializer
        return UserListSerializer


class VehicleViewSet(
    mixins.ListModelMixin,
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    GenericViewSet,
):
    serializer_class = VehicleSerializer

    def get_queryset(self):
        return Vehicle.objects.filter(user_id=self.kwargs['user_pk'])

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx['user_pk'] = self.kwargs.get('user_pk')
        return ctx

    def get_permissions(self):
        if self.action in ('list', 'create'):
            return [IsAuthenticated()]
        return [IsJefeOperacionesOrAbove()]

    def create(self, request, *args, **kwargs):
        target_user_id = str(self.kwargs['user_pk'])
        if str(request.user.id) != target_user_id:
            if not IsJefeOperacionesOrAbove().has_permission(request, self):
                return Response(
                    {'detail': 'No tiene permiso para agregar vehículos a otro usuario.'},
                    status=status.HTTP_403_FORBIDDEN,
                )
        return super().create(request, *args, **kwargs)


class UserImportView(APIView):
    permission_classes = [IsJefeOperacionesOrAbove]

    VALID_ROLES = {r.value for r in Role}
    REQUIRED_COLUMNS = {'codigo_institucional', 'email', 'nombre', 'apellido', 'rol'}

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'Se requiere un archivo.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file)
        except Exception:
            return Response({'detail': 'Archivo Excel inválido.'}, status=status.HTTP_400_BAD_REQUEST)

        ws = wb.active
        headers = [str(cell.value).strip().lower() for cell in ws[1] if cell.value]

        missing = self.REQUIRED_COLUMNS - set(headers)
        if missing:
            return Response(
                {'detail': f'Columnas faltantes: {", ".join(missing)}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created, updated, errors = 0, 0, []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            data = {headers[i]: str(v).strip() if v is not None else '' for i, v in enumerate(row) if i < len(headers)}
            row_errors = self._validate_row(data)

            if row_errors:
                errors.append({'row': row_idx, 'reason': '; '.join(row_errors)})
                continue

            try:
                with transaction.atomic():
                    user, was_created = User.objects.get_or_create(
                        codigo_institucional=data['codigo_institucional'],
                        defaults={
                            'email': data['email'],
                            'nombre': data['nombre'],
                            'apellido': data['apellido'],
                            'rol': data['rol'],
                            'campus_asignado': request.user.campus_asignado,
                        },
                    )
                    if was_created:
                        user.set_password(data['codigo_institucional'])
                        user.save()
                    self._import_vehicles(user, data)
            except Exception as e:
                errors.append({'row': row_idx, 'reason': str(e)})
                continue

            if was_created:
                created += 1
            else:
                updated += 1

        return Response({'created': created, 'updated': updated, 'errors': errors})

    def _validate_row(self, data: dict) -> list:
        row_errors = []
        if not data.get('codigo_institucional'):
            row_errors.append('codigo_institucional requerido')
        if not data.get('email'):
            row_errors.append('email requerido')
        if data.get('rol') not in self.VALID_ROLES:
            row_errors.append(f'rol inválido: {data.get("rol")}')
        return row_errors

    def _import_vehicles(self, user, data: dict):
        for i in ('1', '2'):
            placa = data.get(f'placa_{i}', '').strip()
            tipo = data.get(f'tipo_{i}', '').strip()
            if placa and tipo:
                Vehicle.objects.get_or_create(
                    placa=placa,
                    defaults={'user': user, 'tipo': tipo, 'marca': '', 'modelo': '', 'color': ''},
                )
