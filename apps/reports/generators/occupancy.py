import io
from datetime import date

from django.http import HttpResponse

from apps.access.models import AccessRecord


def _queryset(campus_id: int, date_from: date, date_to: date):
    return AccessRecord.objects.filter(
        campus_id=campus_id,
        entrada_at__date__gte=date_from,
        entrada_at__date__lte=date_to,
    ).select_related('user', 'vehicle', 'space', 'space__lot').order_by('entrada_at')


def _headers():
    return [
        'fecha', 'hora_entrada', 'hora_salida',
        'codigo_usuario', 'nombre_usuario',
        'placa', 'espacio', 'sotano', 'duracion_min',
    ]


def _rows(records):
    rows = []
    for r in records:
        duracion = None
        if r.salida_at:
            duracion = int((r.salida_at - r.entrada_at).total_seconds() / 60)
        rows.append([
            r.entrada_at.date().isoformat(),
            r.entrada_at.strftime('%H:%M'),
            r.salida_at.strftime('%H:%M') if r.salida_at else '',
            r.user.codigo_institucional,
            f'{r.user.nombre} {r.user.apellido}',
            r.vehicle.placa if r.vehicle else '',
            r.space.numero if r.space else '',
            r.space.lot.nombre if r.space else '',
            duracion,
        ])
    return rows


def generate_occupancy_report(campus_id: int, date_from: date, date_to: date, fmt: str) -> HttpResponse:
    records = _queryset(campus_id, date_from, date_to)
    headers = _headers()
    rows = _rows(records)

    filename = f'ocupacion_{date_from}_{date_to}'
    if fmt == 'xlsx':
        return _to_xlsx(headers, rows, filename)
    return _to_pdf(headers, rows, f'Reporte de Ocupación — {date_from} a {date_to}', filename)


def _to_xlsx(headers, rows, filename):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Datos'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F3864')
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row in rows:
        ws.append(row)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


def _to_pdf(headers, rows, title, filename='reporte'):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=20,
    )
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles['Title']), Spacer(1, 12)]

    data = [headers] + rows
    if not rows:
        data = [headers, ['Sin datos para el período seleccionado']]

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F3864')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EBF3FF')]),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)

    response = HttpResponse(buf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    return response
